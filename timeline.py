import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Recreate original "Timeline" sheet content ----------
timeline_data = {
    "Project Months": [
        "Months 1–3 (Apr–Jun 2026)",
        "Months 4–6 (Jul–Sep 2026)",
        "Months 7–9 (Oct–Dec 2026)",
        "Months 10–12 (Jan–Mar 2027)",
        "Months 13–18 (Apr–Sep 2027)",
        "Months 19–24 (Oct 2027–Mar 2028)"
    ],
    "WP1: Bias Evaluation (ROBIMA)": [
        "Design simulation scenarios; pilot runs; develop estimators and models.",
        "Large-scale simulations; evaluate preliminary results.",
        "Finalize simulations; Manuscript 1.",
        "Integrate into ROBIMA; workshop with early adopters.",
        "MSc student N (Qc) visit to Pr. Vo; open-source release; training sessions.",
        "Manuscript 3; updated open-source release."
    ],
    "WP2: New Methods (Individual-Level Data)": [
        "Develop parametric estimators; pilot simulation studies; training on semiparametric NDE/NIE.",
        "Large-scale simulations led by Marie-Félicia; dataset prep by Khoi.",
        "Apply methods to real datasets; refine estimators/models.",
        "Create R package; Manuscript 2.",
        "Apply methods to real datasets; consolidate R package.",
        "Updated open-source release."
    ],
    "WP3: New Methods (Aggregate Data)": [
        "— (WP3 not started).",
        "—",
        "—",
        "—",
        "WP3 starts: design scenarios; parametric estimators; pilot simulations; semiparametric estimators.",
        "Large-scale simulations; real-data applications; estimator validation."
    ],
    "Visits & Dissemination": [
        "JdS 2026 (Clermont-Ferrand, 1–5 Jun): Nicolas | Canada travel MF+Khoi (Jun–Jul)",
        "MSc student N (Qc) visit to Pr. Vo",
        "—",
        "—",
        "Canada travel MF+Khoi (May–Jul 2027) | SSC (May–Jun 2027): MF | ICS Vancouver (20–21 May 2027): MF | Colloque (Qc) | Journées de Biostatistique (Fr)",
        "Present updated outputs at ISCB, JSM"
    ],
    "Manuscripts & Deliverables": [
        "—",
        "—",
        "Manuscript 1 (WP1/WP2)",
        "Manuscript 2 (WP2)",
        "—",
        "Manuscript 3 (WP1/WP2) | Manuscript 4 (WP3)"
    ]
}
df_timeline = pd.DataFrame(timeline_data)

# ---------- Build Month-by-Month sheet with WP3 starting Month 13 ----------
months = list(range(1, 25))

wp1_nicolas = [
    "Kick-off; project organization",
    "Monitoring progress",
    "Coordination meeting; simulation framework design",
    "Progress review; MSc student N visit",
    "Monitor project milestones",
    "Mid-term review",
    "Progress review",
    "Preparation of dissemination plan",
    "Monitor progress",
    "Mid-term evaluation",
    "Monitoring",
    "Review of milestones",
    "Organize training sessions (Québec & France)",
    "Training follow-up",
    "Monitor progress",
    "Continue training sessions",
    "Feedback and review",
    "Finalize mid-term deliverables",
    "Final coordination & reporting",
    "Monitor final progress",
    "Final review",
    "Reporting",
    "Wrap-up & coordination",
    "Final reporting & project closure",
]

wp2_mf = [
    "Design simulation scenarios; parametric estimators",
    "Continue development of estimators; pilot runs",
    "Simulation studies to benchmark estimators",
    "Large-scale simulations (lead); evaluate preliminary results",
    "Continue large-scale simulations",
    "Complete large-scale simulations; preliminary analysis",
    "Finalize simulations; refine estimators",
    "Continue estimator refinement",
    "Complete Manuscript 1 (methods & simulations)",
    "Workshop with early adopters; estimator validation",
    "Continue integration and validation",
    "Estimator validation and integration",
    "Apply new methods to datasets",
    "Continue applications and analyses",
    "Manuscript 2 (applications)",
    "Open-source release (v1)",
    "Application of methods; supervise MSc student",
    "Continue application to datasets",
    "Manuscript 3 (theory & applications)",
    "Continue Manuscript 3",
    "Complete Manuscript 3",
    "Manuscript 3 submission",
    "Dissemination preparations",
    "Ensure all analyses complete",
]

wp2_khoi = [
    "Support pilot simulations",
    "Training on semiparametric estimation of NDE/NIE",
    "Support simulations; initial checks",
    "Support implementation and debugging",
    "Simulation support",
    "Assist with analysis and debugging",
    "Simulation contribution; generate figures/tables",
    "Assist in results generation",
    "Support results tables/figures",
    "Support simulation checks for workshop",
    "Assist with simulation troubleshooting",
    "Support validation",
    "Assist with analysis",
    "Support analyses; results integration",
    "Assist with figures/tables",
    "Support analyses",
    "Support analyses and reproducibility",
    "Assist analyses",
    "Update simulations for reproducibility",
    "Support analyses",
    "Simulation and analysis support",
    "Assist with reproducibility checks",
    "Support dissemination",
    "Final support for reproducibility",
]

wp3_mf_full = [
    "Planning R package integration",
    "Conceptual design of package modules",
    "R package prototype planning",
    "Begin R package structure and modules",
    "R package module development",
    "Package integration planning",
    "Early integration into ROBIMA framework",
    "Package prototype updates",
    "Update package methods",
    "Prototype R package implementation",
    "R package methods refinement",
    "Finalize package prototype",
    "R package documentation & tutorials",
    "Package testing & refinement",
    "Package documentation & tutorials",
    "Update R package based on feedback",
    "R package release (v1)",
    "Finalize package for first release",
    "Updated R package release",
    "Robust R package finalization",
    "R package testing & final updates",
    "R package maintenance",
    "Package final testing",
    "Final R package release",
]

wp3_khoi_full = [
    "Preparation of datasets for training",
    "Dataset cleaning and pipeline practice",
    "Practice application of estimators on sample datasets",
    "Prepare real-world datasets (cleaning, pipelines)",
    "Dataset documentation and formatting",
    "Finalize dataset preparation for analysis",
    "Apply new methods to datasets; contribute to Manuscript 1",
    "Dataset application and checks",
    "Continue real-world application; support Manuscript 1",
    "Dataset expansion; preprocessing functions",
    "Apply preprocessing on datasets; integration into package",
    "Continue data prep for package integration",
    "Support application integration; participate in training",
    "Dataset preparation and integration",
    "Support applications and package tutorials",
    "Dataset integration into package; training support",
    "Continue dataset support; training participation",
    "Dataset integration and maintenance",
    "Contribute to Manuscript 3; maintain datasets",
    "Dataset curation; applied results contribution",
    "Final dataset integration",
    "Contribute to final results",
    "Final dataset checks",
    "Dataset handover and documentation",
]

# Shift WP3 to start at Month 13 (first 12 months blank; months 13-24 use the first 12 planned tasks)
wp3_mf_shifted = ["—"] * 12 + wp3_mf_full[:12]
wp3_khoi_shifted = ["—"] * 12 + wp3_khoi_full[:12]

# Visits & Dissemination per month (Month 1 = Apr 2026)
visits = [""] * 24
manuscripts_m = [""] * 24

# JdS 2026 (1–5 Jun 2026) -> Month 3
visits[2] = "JdS 2026 (Clermont-Ferrand, 1–5 Jun): Nicolas | Canada travel MF+Khoi (Jun–Jul)"
# Canada travel Jun–Jul 2026 -> Months 3–4
visits[3] = "Canada travel MF+Khoi (Jun–Jul) | MSc student N visit Pr. Vo"
# MSc student N visit Jul–Sep 2026 -> Months 4–6
visits[4] = "MSc student N visit Pr. Vo"
visits[5] = "MSc student N visit Pr. Vo"

# SSC Annual Meeting and ICS in 2027, plus Canada travel May–Jul 2027 -> Months 14–16
visits[13] = "SSC (May–Jun 2027): MF | ICS Vancouver (20–21 May 2027): MF | Canada travel MF+Khoi (May–Jul)"
visits[14] = "Canada travel MF+Khoi (May–Jul) | Colloque francophone interfacultaire (Qc)"
visits[15] = "Canada travel MF+Khoi (May–Jul) | Journées de Biostatistique (Fr)"

# Manuscripts
manuscripts_m[8]  = "Manuscript 1 (methods & simulations)"
manuscripts_m[14] = "Manuscript 2 (applications)"
manuscripts_m[18] = "Manuscript 3 – draft"
manuscripts_m[20] = "Manuscript 3 – complete"
manuscripts_m[21] = "Manuscript 3 – submission"

df_month = pd.DataFrame({
    "Month": months,
    "WP1 – Nicolas": wp1_nicolas,
    "WP2 – MF": wp2_mf,
    "WP2 – Khoi": wp2_khoi,
    "WP3 – MF": wp3_mf_shifted,
    "WP3 – Khoi": wp3_khoi_shifted,
    "Visits & Dissemination": visits,
    "Manuscripts & Deliverables": manuscripts_m
})

# ---------- Write workbook ----------
out_path = "WP_Timeline_py.xlsx"
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Timeline sheet
    df_timeline.to_excel(writer, index=False, sheet_name="Timeline")
    ws_t = writer.sheets["Timeline"]
    # Month-by-Month sheet
    df_month.to_excel(writer, index=False, sheet_name="Month-by-Month")
    ws_m = writer.sheets["Month-by-Month"]
    
    # Styling function
    def style_sheet(ws):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        wrap = Alignment(wrap_text=True, vertical="top")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Cells
        col_widths = {i+1: 12 for i in range(ws.max_column)}
        for r in range(2, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = wrap
                cell.border = border
                val = str(cell.value) if cell.value is not None else ""
                col_widths[c] = min(max(col_widths[c], len(val) + 2), 55)
        for c in range(1, ws.max_column+1):
            ws.column_dimensions[get_column_letter(c)].width = col_widths[c]
        ws.freeze_panes = "A2"
    
    style_sheet(ws_t)
    style_sheet(ws_m)

out_path
